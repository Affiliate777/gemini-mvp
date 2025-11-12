from pathlib import Path
from openpyxl import load_workbook

def parse_excel(path: Path):
    """Return a canonical dict for an excel file.
    For MVP: read first sheet, take first row as headers and second row as values.
    """
    wb = load_workbook(filename=str(path), data_only=True)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty sheet")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    first_data = rows[1] if len(rows) > 1 else [None]*len(headers)
    fields = {}
    for k, v in zip(headers, first_data):
        if k:
            fields[k] = v
    doc = {
        "type": "excel",
        "path": str(path),
        "sheet": sheet.title,
        "metadata": {"nrows": len(rows), "ncols": len(headers)},
        "fields": fields
    }
    return doc
