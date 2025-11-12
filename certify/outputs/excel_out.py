from pathlib import Path
import pandas as pd
from typing import List, Dict
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def _flatten_results(results: List[Dict]) -> List[Dict]:
    rows = []
    for r in results:
        file = r.get("file")
        checks = r.get("checks")
        if isinstance(checks, dict):
            flat = checks.get("simple", []) + checks.get("rules", [])
        else:
            flat = checks or []
        for c in flat:
            rows.append({
                "file": file,
                "rule_id": c.get("id"),
                "title": c.get("title"),
                "severity": (c.get("severity") or "").lower(),
                "passed": c.get("passed"),
                "info": c.get("info")
            })
    return rows

def write_excel_summary(results: List[Dict], out_path: Path):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Summary numbers
    rows = _flatten_results(results)
    total_files = len(results)
    total_checks = len(rows)
    passed = sum(1 for r in rows if r.get("passed"))
    warnings = sum(1 for r in rows if r.get("severity") == "warning" and not r.get("passed"))
    errors = sum(1 for r in rows if r.get("severity") == "error" and not r.get("passed"))

    summary = {
        "total_files": [total_files],
        "total_checks": [total_checks],
        "passed": [passed],
        "warnings": [warnings],
        "errors": [errors]
    }
    df_summary = pd.DataFrame(summary)

    # Details sheet
    if rows:
        df_details = pd.DataFrame(rows)
    else:
        df_details = pd.DataFrame(columns=["file", "rule_id", "title", "severity", "passed", "info"])

    # Write excel workbook
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_details.to_excel(writer, sheet_name="Checks", index=False)

    # Apply highlighting using openpyxl
    wb = load_workbook(out_path)
    ws = wb["Checks"]

    # define fills
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="FFDFF0D8", end_color="FFDFF0D8", fill_type="solid")

    # find column indexes by header row (row 1)
    headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

    sev_col = headers.get("severity")
    passed_col = headers.get("passed")
    if sev_col is None or passed_col is None:
        # nothing to do
        wb.save(out_path)
        return out_path

    max_row = ws.max_row
    max_col = ws.max_column

    # apply fills row-by-row based on severity and passed
    for row in range(2, max_row+1):
        sev = ws.cell(row=row, column=sev_col).value
        passed_val = ws.cell(row=row, column=passed_col).value
        sev_norm = (str(sev).lower() if sev is not None else "")
        # choose fill: error -> red (if not passed), warning -> yellow (if not passed), else green if passed
        if sev_norm == "error" and not passed_val:
            fill = red_fill
        elif sev_norm == "warning" and not passed_val:
            fill = yellow_fill
        elif passed_val:
            fill = green_fill
        else:
            fill = None
        if fill:
            for col in range(1, max_col+1):
                ws.cell(row=row, column=col).fill = fill

    # auto-adjust column widths modestly
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value)
            except Exception:
                val = ""
            if val:
                max_length = max(max_length, len(val))
        adjusted = (max_length + 2) if max_length < 60 else 60
        ws.column_dimensions[col_letter].width = adjusted

    wb.save(out_path)
    return out_path
